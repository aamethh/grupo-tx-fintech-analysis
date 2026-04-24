"""
export_excel.py
---------------
Produces a multi-sheet Excel workbook for executive distribution.
  Sheet 1: Summary          — KPI snapshot per company / year
  Sheet 2: Financials       — full balance sheet + P&L
  Sheet 3: Ratios           — all calculated metrics (%)
  Sheet 4: AUM Trend        — monthly time series
  Sheet 5: Peer Comparison  — side-by-side benchmarking
"""
import pandas as pd
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from sqlalchemy import text
from db_connect import get_engine

OUT = "outputs/grupo_tx_executive_report.xlsx"

NAVY  = "1B3A6B"
GOLD  = "C8972B"
WHITE = "FFFFFF"
LGRAY = "F3F4F6"


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_rows(ws, start_row=2):
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start=0):
        fill = PatternFill("solid", fgColor=LGRAY) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)


def highlight_kpi(ws, col_idx, threshold, good_color="C6EFCE", bad_color="FFC7CE"):
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                color = good_color if cell.value >= threshold else bad_color
                cell.fill = PatternFill("solid", fgColor=color)


def fetch(engine, query):
    with engine.connect() as conn:
        return pd.read_sql(text(query), conn)


def build_excel():
    engine = get_engine()

    financials_q = """
        SELECT c.company_code AS Company, f.fiscal_year AS Year,
               f.total_assets/1e6 AS Assets_M, f.gross_loans/1e6 AS Loans_M,
               f.total_deposits/1e6 AS Deposits_M, f.total_equity/1e6 AS Equity_M,
               f.aum/1e6 AS AUM_M, f.net_income/1e6 AS NetIncome_M,
               f.net_revenue/1e6 AS Revenue_M, f.fee_income/1e6 AS FeeIncome_M,
               f.capital_ratio*100 AS CapitalRatio, f.npl_coverage*100 AS NPL_Coverage
        FROM financials f JOIN companies c ON c.company_id=f.company_id
        ORDER BY c.company_code, f.fiscal_year"""

    metrics_q = """
        SELECT c.company_code AS Company, m.fiscal_year AS Year,
               ROUND(m.roe*100,2) AS ROE_Pct, ROUND(m.roa*100,2) AS ROA_Pct,
               ROUND(m.cost_to_income*100,2) AS CostToIncome_Pct,
               ROUND(m.loan_to_deposit*100,2) AS LoanToDeposit_Pct,
               ROUND(m.aum_to_assets*100,2) AS AUM_Assets_Pct,
               ROUND(m.capital_multiplier,2) AS CapitalMult,
               ROUND(m.aum_growth_yoy*100,2) AS AUM_Growth_Pct,
               ROUND(m.asset_growth_yoy*100,2) AS Asset_Growth_Pct
        FROM metrics m JOIN companies c ON c.company_id=m.company_id
        ORDER BY c.company_code, m.fiscal_year"""

    timeseries_q = """
        SELECT c.company_code AS Company,
               FORMAT(ts.period_date,'MMM yyyy') AS Period,
               ts.metric_value/1e6 AS AUM_M
        FROM time_series ts JOIN companies c ON c.company_id=ts.company_id
        WHERE ts.metric_name='aum'
        ORDER BY ts.company_id, ts.period_date"""

    peer_q = """
        SELECT c.company_code AS Company, f.fiscal_year AS Year,
               ROUND(f.net_income/NULLIF(f.total_equity,0)*100,2) AS ROE,
               ROUND(f.net_income/NULLIF(f.total_assets,0)*100,2) AS ROA,
               ROUND(f.aum/NULLIF(f.total_assets,0)*100,2) AS AUM_Penetration,
               f.capital_ratio*100 AS CapitalRatio, f.npl_coverage*100 AS NPL_Coverage
        FROM financials f JOIN companies c ON c.company_id=f.company_id
        ORDER BY f.fiscal_year, ROE DESC"""

    df_fin  = fetch(engine, financials_q)
    df_met  = fetch(engine, metrics_q)
    df_ts   = fetch(engine, timeseries_q)
    df_peer = fetch(engine, peer_q)

    summary = df_met[df_met["Year"] == df_met["Year"].max()].copy()

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        for df, sheet in [
            (summary,  "Summary"),
            (df_fin,   "Financials"),
            (df_met,   "Ratios"),
            (df_ts,    "AUM Trend"),
            (df_peer,  "Peer Comparison"),
        ]:
            df.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(OUT)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        style_header(ws)
        style_rows(ws)
        autofit(ws)
        ws.freeze_panes = "A2"

    # Highlight ROE column in Ratios sheet
    ws_r = wb["Ratios"]
    roe_col = [cell.column for cell in ws_r[1] if cell.value == "ROE_Pct"]
    if roe_col:
        highlight_kpi(ws_r, roe_col[0], threshold=15.0)

    wb.save(OUT)
    print(f"Excel exported -> {OUT}")


if __name__ == "__main__":
    build_excel()
